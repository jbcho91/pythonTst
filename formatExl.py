from openpyxl import load_workbook

# 변경할 엑셀 파일의 경로를 저장
load_wb = load_workbook("", data_only=True)
ws = load_wb.active

i = ws.max_row
j = 0
print(str(i) + "행 존재")

while i >= 2:
    print(str(i) + '번째 행')
    # 10번째 행이 None 값을 가진 row 들을 선별 및 삭제
    if ws.cell(i, 10).value is None:
        ws.delete_rows(i)
        j = j+1
        print(str(j) + '번째 삭제')
    else:
        # 삭제할 row 가 아니면 7번째 열에 /가 있는지 판별 해서 / 삭제
        cell_value = ws.cell(i, 7).value
        if cell_value is not None:
            if '/' in cell_value:
                cell_value = cell_value.replace('/', '')
                ws.cell(i, 7).value = cell_value
    i -= 1

print(str(j) + '개 삭제')
# 새로 바뀐 데이터 저장
load_wb.save("")


