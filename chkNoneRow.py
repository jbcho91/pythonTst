from openpyxl import load_workbook

load_wb = load_workbook("", data_only=True)
load_ws = load_wb['Sheet']

j = 0
print("check None Row"+str(load_ws.max_row))

for i in range(2, load_ws.max_row):
    if load_ws.cell(i, 10).value is None:
        print(str(i)+'번째 행')
        print(load_ws.cell(i, 6).value)
        print(load_ws.cell(i, 7).value)
        print(load_ws.cell(i, 8).value)
        print(load_ws.cell(i, 9).value)
        print(load_ws.cell(i, 10).value)
        j = j + 1

print("End" + str(j) + "개 존재")



# 셀 주소로 값 출력
#print(load_ws['B1'].value)



